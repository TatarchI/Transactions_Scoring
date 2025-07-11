"""
📦 Module6_compare_isoforest_multi.py
Author: Ivan Tatarchuk

Description:
- Compares the results of IsoForest and Multi-Criteria anomaly scoring models.
- Outputs key matching and divergence metrics for interpretability.
- Calculates confidence and structured comments for stakeholder review.
- Builds a **rank-based combined final scoring** for decision support.
- Saves a final, visually highlighted Excel table for transparent anomaly review.
"""

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ----------------------------------------
# 📌 Paths
# ----------------------------------------
isoforest_path = 'check_files/isoforest/scoring_output_isoforest.csv.gz'
multi_path = 'check_files/multi_criteria/scoring_output_multi_criteria.csv.gz'
output_dir = 'Final_decision'
os.makedirs(output_dir, exist_ok=True)

# ----------------------------------------
# 📌 Load IsoForest and Multi-Criteria results
# ----------------------------------------
print("✅ Loading IsoForest results...")
df_iso = pd.read_csv(isoforest_path, compression='gzip')

print("✅ Loading Multi-Criteria results...")
df_multi = pd.read_csv(multi_path, compression='gzip')

# ----------------------------------------
# 📌 Merge results for comparison
# ----------------------------------------
df_compare = df_iso.merge(df_multi, on='ID', suffixes=('_isoforest', '_multi'))

# ----------------------------------------
# 📌 Overlap analysis
# ----------------------------------------
same_outlier = (df_compare['is_outlier_isoforest'] == df_compare['is_outlier_multi']).mean()
both_anomaly = ((df_compare['is_outlier_isoforest'] == 1) & (df_compare['is_outlier_multi'] == 1)).sum()
both_normal = ((df_compare['is_outlier_isoforest'] == 0) & (df_compare['is_outlier_multi'] == 0)).sum()
iso_only = ((df_compare['is_outlier_isoforest'] == 1) & (df_compare['is_outlier_multi'] == 0)).sum()
multi_only = ((df_compare['is_outlier_isoforest'] == 0) & (df_compare['is_outlier_multi'] == 1)).sum()

print("\n📊 Overlap and divergence summary:")
print(f"🔹 Both detect anomaly: {both_anomaly}")
print(f"🔹 Both detect normal: {both_normal}")
print(f"🔹 IsoForest-only anomaly: {iso_only}")
print(f"🔹 Multi-Criteria-only anomaly: {multi_only}")

# ----------------------------------------
# 📌 Confidence and comment calculation
# ----------------------------------------

def determine_confidence(row):
    """
    Determines confidence score:
    - 1 if both models detect anomaly.
    - 0.5 if one model detects anomaly.
    - 0 if both detect normal.
    """
    if row['is_outlier_isoforest'] == 1 and row['is_outlier_multi'] == 1:
        return 1
    elif row['is_outlier_isoforest'] == 1 or row['is_outlier_multi'] == 1:
        return 0.5
    else:
        return 0

def determine_comment(row):
    """
    Generates a readable comment for stakeholder reporting:
    - 'both', 'isoforest_only', 'multi_only', 'none'
    """
    if row['is_outlier_isoforest'] == 1 and row['is_outlier_multi'] == 1:
        return 'both'
    elif row['is_outlier_isoforest'] == 1:
        return 'isoforest_only'
    elif row['is_outlier_multi'] == 1:
        return 'multi_only'
    else:
        return 'none'

def determine_system_decision(row):
    """
    Generates a clear system decision for business review:
    - 'Yes' (requires investigation)
    - 'Maybe Yes' (potentially suspicious)
    - 'No' (likely normal)
    """
    if row['confidence'] == 1:
        return 'Yes'
    elif row['confidence'] == 0.5:
        return 'Maybe Yes'
    else:
        return 'No'

df_compare['confidence'] = df_compare.apply(determine_confidence, axis=1)
df_compare['comment'] = df_compare.apply(determine_comment, axis=1)
df_compare['system_decision'] = df_compare.apply(determine_system_decision, axis=1)

# ----------------------------------------
# 📌 Ranking for final scoring
# ----------------------------------------
df_compare['rank_isoforest'] = df_compare['anomaly_score'].rank(method='min', ascending=True)
df_compare['rank_multi'] = df_compare['multi_criteria_score'].rank(method='min', ascending=False)
df_compare['final_rank_score'] = df_compare['rank_isoforest'] + df_compare['rank_multi']

# ----------------------------------------
# 📌 Final column order
# ----------------------------------------
final_columns = [
    'ID', 'system_decision', 'final_rank_score',
    'anomaly_score', 'multi_criteria_score',
    'is_outlier_isoforest', 'is_outlier_multi',
    'confidence', 'comment',
    'rank_isoforest', 'rank_multi'
]
# multi_criteria_score - the higher the score, the higher the anomaly likelihood
# anomaly_score - the lower the score, the higher the anomaly likelihood
df_final = df_compare[final_columns]

# ----------------------------------------
# 📌 Save final Excel with highlighted columns
# ----------------------------------------
final_path = f'{output_dir}/final_scoring_decision.xlsx'
df_final.to_excel(final_path, index=False)

wb = load_workbook(final_path)
ws = wb.active

# Highlight key columns
fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
header_row = 1
for col_idx, col_name in enumerate(df_final.columns, 1):
    if col_name in ['system_decision', 'final_rank_score']:
        ws.cell(row=header_row, column=col_idx).fill = fill

wb.save(final_path)

print(f"\n✅ Final decision table saved to {final_path}")
print("\n✅ Module completed.")

"""
🔹 Module 6 Summary (Module6_compare_isoforest_multi.py):

This module finalizes our pipeline by **comparing IsoForest and Multi-Criteria anomaly detection models**, generating:
- Clean, stakeholder-friendly columns: `system_decision`, `confidence`, `final_rank_score`.
- Cluster categorization: `Yes` (requires review), `Maybe Yes` (potential risk), `No` (normal).
- Rank-based fusion of IsoForest (low score = anomaly) and Multi-Criteria (high score = anomaly) for robust prioritization.
- A final Excel file with highlights for immediate review and integration into business processes.

The approach ensures that:
✅ Consistent, explainable decisions are generated for each transaction.
✅ Business stakeholders receive an actionable, interpretable list with risk stratification.
✅ The system can be adjusted for different sensitivity levels (2-3% anomaly thresholds) depending on project needs.

---
🔹 Project-Wide Final Summary:

This treasury transaction scoring pipeline demonstrates:
- **Robust preprocessing**, outlier handling, and scalable feature engineering across 730K+ transactions.
- Dual anomaly detection via IsoForest (machine learning-based) and Multi-Criteria (transparent, explainable model).
- Fusion of scores with clear clustering (`Yes`, `Maybe Yes`, `No`) to enable targeted fraud checks and process optimizations.
- A fully automated, reproducible pipeline producing high-signal anomaly candidates for manual review.
- Business Value:
    • Reduces investigation workload by prioritizing high-risk transactions.
    • Provides explainable risk drivers via Multi-Criteria contribution analysis.
    • Supports audits, compliance, and fraud detection initiatives transparently.
    
🔹 Why these two models were chosen:

We selected **Isolation Forest and Multi-Criteria Scoring** because they **complement each other.
Isolation Forest efficiently handles large-scale transactional data without requiring labels, delivers high silhouette 
scores, and reliably detects outliers that rule-based methods might miss. However, it remains a black-box model, making
it challenging to explain why a specific transaction was flagged as anomalous. To address interpretability, we added a
manual Multi-Criteria model using robust normalization and feature weights, providing clear scoring and immediate 
visibility into which features contributed most to a transaction’s risk. This transparency is crucial for business,
audit, and explainability to stakeholders. Together, these two models deliver robust and interpretable anomaly scoring, 
enabling effective segmentation into *Yes, Maybe Yes, No* clusters and allowing the team to focus on the highest-risk 
transactions while maintaining transparency for compliance and management reporting.

"""